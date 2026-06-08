import logging
from contextlib import contextmanager
import openpyxl
import pytest
from pathlib import Path
from fosslight_scanner import fosslight_scanner as scanner_module
from fosslight_scanner.fosslight_scanner import run_scanner, download_source, init, run_main, run_dependency
from fosslight_util import constant as fosslight_constant
import fosslight_util.set_log as set_log_module
from fosslight_util.oss_item import ScannerItem
from fosslight_util.constant import FOSSLIGHT_BINARY, FOSSLIGHT_DEPENDENCY, FOSSLIGHT_SOURCE, SHEET_NAME_FOR_SCANNER

_DEFAULT_OUTPUT_DIR = scanner_module._output_dir
_DEFAULT_LOG_FILE = scanner_module._log_file

_SRC = SHEET_NAME_FOR_SCANNER[FOSSLIGHT_SOURCE]       # 'SRC_FL_Source'
_BIN = SHEET_NAME_FOR_SCANNER[FOSSLIGHT_BINARY]       # 'BIN_FL_Binary'
_DEP = SHEET_NAME_FOR_SCANNER[FOSSLIGHT_DEPENDENCY]   # 'DEP_FL_Dependency'
_TESTS_DIR = Path(__file__).resolve().parent
_SCAN_PROJECT_DIR = _TESTS_DIR / "fixtures" / "scan_project"
_TEST_CAPTURE_LOG = "pytest_run_main_capture.txt"
_MIN_DATA_ROWS = 2

SHEET_CHECK_PARAMS = [
    pytest.param(["all"],        [_SRC, _BIN, _DEP], id="all"),
    pytest.param(["source"],     [_SRC],             id="source"),
    pytest.param(["binary"],     [_BIN],             id="binary"),
    pytest.param(["dependency"], [_DEP],             id="dependency"),
]


def _get_sheet_names(xlsx_path: str) -> list:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _get_sheet_row_count(xlsx_path: str, sheet_name: str) -> int:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    row_count = ws.max_row or 0
    wb.close()
    return row_count


def _reset_fosslight_logger() -> None:
    logger = logging.getLogger(fosslight_constant.LOGGER_NAME)
    for handler in logger.handlers[:]:
        handler.close()
        logger.removeHandler(handler)


def _test_capture_log_path(output_dir: Path) -> Path:
    return output_dir / "fosslight_log" / _TEST_CAPTURE_LOG


def _format_log_for_failure(content: str, max_chars: int = 12000) -> str:
    if len(content) <= max_chars:
        return content
    keywords = ("ERROR", "Failed", "FOSSLIGHT_SOURCE", "scancode", "Detected", "Scan Result")
    highlights = [line for line in content.splitlines() if any(key in line for key in keywords)]
    tail = content.splitlines()[-100:]
    parts = ["...(log truncated)..."]
    if highlights:
        parts.append("--- highlights ---")
        parts.extend(highlights[-50:])
    parts.append("--- tail ---")
    parts.extend(tail)
    return "\n".join(parts)[:max_chars]


def _collect_log_search_dirs(output_dir: Path) -> list:
    search_dirs = [
        output_dir / "fosslight_log",
        _TESTS_DIR / "fosslight_log",
        Path.cwd() / "fosslight_log",
    ]
    for base in (output_dir.parent, _TESTS_DIR, Path.cwd()):
        for temp_dir in base.glob(".fosslight_temp_*"):
            search_dirs.append(temp_dir / "fosslight_log")
    return search_dirs


def _snapshot_fosslight_logger():
    logger = logging.getLogger(fosslight_constant.LOGGER_NAME)
    return {
        "handlers": logger.handlers[:],
        "level": logger.level,
        "propagate": logger.propagate,
        "disabled": logger.disabled,
    }


def _restore_fosslight_logger(snapshot: dict) -> None:
    logger = logging.getLogger(fosslight_constant.LOGGER_NAME)
    for handler in logger.handlers[:]:
        handler.flush()
        handler.close()
        logger.removeHandler(handler)
    logger.setLevel(snapshot["level"])
    logger.propagate = snapshot["propagate"]
    logger.disabled = snapshot["disabled"]
    for handler in snapshot["handlers"]:
        logger.addHandler(handler)


@contextmanager
def _run_main_log_capture(output_dir: Path):
    capture_log_path = _test_capture_log_path(output_dir)
    capture_log_path.parent.mkdir(parents=True, exist_ok=True)

    base_logger = logging.getLogger(fosslight_constant.LOGGER_NAME)
    logger_snapshot = _snapshot_fosslight_logger()
    for handler in logger_snapshot["handlers"]:
        base_logger.removeHandler(handler)

    original_init_log = set_log_module.init_log
    capture_handler = logging.FileHandler(capture_log_path, encoding="utf-8")
    capture_handler.setLevel(logging.DEBUG)
    capture_handler.setFormatter(logging.Formatter("[%(levelname)7s] %(message)s"))

    def patched_init_log(log_file, *args, **kwargs):
        _reset_fosslight_logger()
        logger, result_log = original_init_log(log_file, *args, **kwargs)
        if capture_handler not in base_logger.handlers:
            base_logger.addHandler(capture_handler)
        return logger, result_log

    set_log_module.init_log = patched_init_log
    scanner_module.init_log = patched_init_log
    _reset_fosslight_logger()
    try:
        yield capture_log_path
    finally:
        set_log_module.init_log = original_init_log
        scanner_module.init_log = original_init_log
        _restore_fosslight_logger(logger_snapshot)


def _read_run_main_log(output_dir: Path) -> str:
    capture_log_path = _test_capture_log_path(output_dir)
    if capture_log_path.is_file():
        try:
            content = capture_log_path.read_text(encoding="utf-8", errors="replace")
        except OSError as err:
            return f"(failed to read {capture_log_path}: {err})"
        if content.strip():
            return f"(from {capture_log_path})\n{_format_log_for_failure(content)}"

    log_files = []
    searched = []
    for log_dir in _collect_log_search_dirs(output_dir):
        searched.append(str(log_dir))
        if log_dir.is_dir():
            log_files.extend(log_dir.glob("*.txt"))
    if not log_files:
        return f"(no log file; searched: {', '.join(searched)})"

    log_files_with_mtime = []
    for path in log_files:
        try:
            log_files_with_mtime.append((path.stat().st_mtime, path))
        except OSError:
            continue

    if not log_files_with_mtime:
        return f"(no readable log file; searched: {', '.join(searched)})"

    log_files_with_mtime.sort(key=lambda item: item[0], reverse=True)
    _, latest = log_files_with_mtime[0]

    try:
        content = latest.read_text(encoding="utf-8", errors="replace")
    except OSError as err:
        return f"(failed to read {latest}: {err})"

    return f"(from {latest})\n{_format_log_for_failure(content)}"


def _fail_with_run_main_log(output_dir: Path, message: str) -> None:
    pytest.fail(f"{message}\n\n--- run_main log ---\n{_read_run_main_log(output_dir)}")


def test_run_dependency(tmp_path):
    # given
    path_to_analyze = tmp_path / "test_project"
    output_file_with_path = tmp_path / "output_file"
    params = "-m 'npm' -a 'activate_cmd' -d 'deactivate_cmd' -c 'custom_dir' -n 'app_name' -t 'token_value'"
    path_to_exclude = ["node_modules"]

    # Create the directory to analyze
    path_to_analyze.mkdir(parents=True, exist_ok=True)

    # when
    result = run_dependency(str(path_to_analyze), str(output_file_with_path), params, path_to_exclude)

    # then
    # Check that result is an instance of ScannerItem
    assert isinstance(result, ScannerItem), "Result should be an instance of ScannerItem."

    # Check that result is not None
    assert result is not None, "Result should not be None."


def test_run_scanner(tmp_path):
    # given
    src_path = tmp_path / "test_src"
    output_path = tmp_path / "output"
    dep_arguments = ""
    output_file = ['test_output']
    output_extension = [".yaml"]

    # Create necessary directories and files for the test
    src_path.mkdir(parents=True, exist_ok=True)
    output_path.mkdir(parents=True, exist_ok=True)

    # Create a dummy source file in `src_path` to be scanned
    test_file = src_path / "test_file.py"
    test_file.write_text("# This is a test file\nprint('Hello, World!')")

    # when
    run_scanner(
        src_path=str(src_path),
        dep_arguments=dep_arguments,
        output_path=str(output_path),
        keep_raw_data=True,
        run_src=True,
        run_bin=False,
        run_dep=False,
        run_prechecker=False,
        remove_src_data=False,
        result_log={},
        output_files=output_file,
        output_extensions=output_extension,
        num_cores=1,
        path_to_exclude=[]
    )

    # then
    # Check if `src_path` and `output_path` still exist
    assert src_path.is_dir(), "Source path should still exist."
    assert output_path.is_dir(), "Output path should still exist."


def test_download_source(tmp_path):
    # given
    link = "https://example.com/test_repo.git"
    out_dir = tmp_path / "output"

    # Create the necessary output directory
    out_dir.mkdir(parents=True, exist_ok=True)

    # when
    success, temp_src_dir, oss_name, oss_version = download_source(str(link), str(out_dir))

    # then
    # Ensure the function completes successfully
    assert isinstance(success, bool), "Function should return a boolean for success."

    # If the function fails, ensure temp_src_dir is empty
    if not success:
        assert temp_src_dir == "", "temp_src_dir should be an empty string if download fails."
    else:
        # If the function succeeds, check temp_src_dir is a valid path
        assert isinstance(temp_src_dir, str), "Temporary source directory should be a string."
        assert str(temp_src_dir).startswith(str(out_dir)), "Temporary source directory should be within the output directory."

    # Ensure oss_name and oss_version are strings
    assert isinstance(oss_name, str), "OSS name should be a string."
    assert isinstance(oss_version, str), "OSS version should be a string."


def test_init(tmp_path):
    # given
    output_path = tmp_path / "output"
    scanner_module._output_dir = "test_output"
    scanner_module._log_file = "test_log"

    try:
        # when
        dir_created, output_root_dir, result_log = init(str(output_path))

        # then
        assert dir_created is True, "The output directory should be created."
        assert output_root_dir == str(output_path), "Output root directory should match the given path."
        assert result_log is not None, "Result log should not be None."
        assert isinstance(result_log, dict), "Result log should be a dictionary."
    finally:
        scanner_module._output_dir = _DEFAULT_OUTPUT_DIR
        scanner_module._log_file = _DEFAULT_LOG_FILE
        _reset_fosslight_logger()


def test_run_main(tmp_path):
    # given
    mode_list = ["source"]
    path_arg = [str(tmp_path / "test_src")]
    dep_arguments = []
    output_file_or_dir = str(tmp_path / "output")
    file_format = ['yaml']
    url_to_analyze = ""
    db_url = ""

    # Create necessary directories and files for the test
    (tmp_path / "test_src").mkdir(parents=True, exist_ok=True)

    # when
    result = run_main(
        mode_list=mode_list,
        path_arg=path_arg,
        dep_arguments=dep_arguments,
        output_file_or_dir=output_file_or_dir,
        file_format=file_format,
        url_to_analyze=url_to_analyze,
        db_url=db_url,
        hide_progressbar=True,  # Disable progress bar for testing
        keep_raw_data=True,     # Keep raw data to avoid cleanup during test
        num_cores=1,
        correct_mode=True,
        correct_fpath="",
        ui_mode=False,
        path_to_exclude=[]
    )

    # then
    assert result is True


@pytest.mark.parametrize("mode_list,expected_sheets", SHEET_CHECK_PARAMS)
def test_output_excel_contains_required_sheets(tmp_path, mode_list, expected_sheets):
    # given
    output_dir = tmp_path / "output"

    # when
    with _run_main_log_capture(output_dir):
        result = run_main(
            mode_list=mode_list,
            path_arg=[str(_SCAN_PROJECT_DIR)],
            dep_arguments=[],
            output_file_or_dir=str(output_dir),
            file_format=["excel"],
            url_to_analyze="",
            db_url="",
            hide_progressbar=True,
            keep_raw_data=False,
            num_cores=1,
            correct_mode=False,
            correct_fpath="",
            ui_mode=False,
            path_to_exclude=[]
        )

    # then
    if not result:
        _fail_with_run_main_log(output_dir, "run_main should return True on success.")

    xlsx_files = list(output_dir.glob("*.xlsx"))
    if len(xlsx_files) != 1:
        _fail_with_run_main_log(
            output_dir,
            f"Expected exactly one Excel report in {output_dir}, got {len(xlsx_files)}: {xlsx_files}",
        )

    xlsx_file = str(xlsx_files[0])

    actual_sheets = _get_sheet_names(xlsx_file)
    for sheet in expected_sheets:
        if sheet not in actual_sheets:
            _fail_with_run_main_log(
                output_dir,
                f"[mode={mode_list}] Sheet '{sheet}' not found. "
                f"Actual sheets: {actual_sheets}. "
                "One of the scanners may have failed silently.",
            )

    for sheet in expected_sheets:
        row_count = _get_sheet_row_count(xlsx_file, sheet)
        if row_count < _MIN_DATA_ROWS:
            _fail_with_run_main_log(
                output_dir,
                f"[mode={mode_list}] Sheet '{sheet}' has {row_count} row(s), "
                f"expected at least {_MIN_DATA_ROWS}.",
            )
